"""XLSX parser（openpyxl）—— 抽各 sheet 数据为 csv-like 文本

每个 sheet：标题 → 行内 cells 用 ' | ' 连接；空行跳过。整工作簿用 \n\n 串。
"""

from __future__ import annotations

import io
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from chameleon.api.knowledge.parsers import ParsedDocument


async def parse(source: bytes | str, *, name: str) -> "ParsedDocument":
    from chameleon.api.knowledge.parsers import ParsedDocument

    if isinstance(source, str):
        raise TypeError("xlsx parser requires bytes input")

    try:
        from openpyxl import load_workbook
    except ImportError as e:  # noqa: BLE001
        raise RuntimeError("openpyxl not installed; uv add openpyxl") from e

    wb = load_workbook(filename=io.BytesIO(source), data_only=True, read_only=True)
    sheets_text: list[str] = []
    for ws in wb.worksheets:
        parts: list[str] = [f"--- Sheet: {ws.title} ---"]
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                parts.append(" | ".join(cells))
        sheets_text.append("\n".join(parts))
    wb.close()

    text = "\n\n".join(sheets_text)
    metadata: dict = {"name": name, "sheets": [s.title for s in wb.worksheets]}
    return ParsedDocument(text=text, metadata=metadata)
